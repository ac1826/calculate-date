from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from processor import export_summary_workbook, load_customer_mapping, process_files, transform_dalian


def to_excel_bytes(frame: pd.DataFrame) -> BytesIO:
    output = BytesIO()
    frame.to_excel(output, index=False)
    output.seek(0)
    return output


def sample_dalian() -> BytesIO:
    return to_excel_bytes(
        pd.DataFrame(
            [
                {
                    "销售凭证": "10109593",
                    "售达方": "GBK0000",
                    "售达方.1": "新加坡CMM",
                    "描述": "新加坡CMM 脆皮鸡肉1000g/12袋/箱",
                    "订购数量": 2400,
                    "订购数量.1": "KG",
                    "单价PN00": 3.88,
                    "单价PN00.1": "USD",
                },
                {
                    "销售凭证": "10109594",
                    "售达方": "GCV0000",
                    "售达方.1": "DOUBLE CHIN FOOD SERVICES PTE LTD",
                    "描述": "新加坡DC照烧鸡排100g/10枚/12袋/箱",
                    "订购数量": 4800,
                    "订购数量.1": "KG",
                    "单价PN00": 4.65,
                    "单价PN00.1": "USD",
                },
                {
                    "销售凭证": "60196329",
                    "售达方": "GCV0000",
                    "售达方.1": "DOUBLE CHIN FOOD SERVICES PTE LTD",
                    "描述": "新加坡DC照烧鸡排100g/10枚/12袋/箱",
                    "订购数量": -3240,
                    "订购数量.1": "KG",
                    "单价PN00": 4.65,
                    "单价PN00.1": "USD",
                },
            ]
        )
    )


def sample_tieling() -> BytesIO:
    return to_excel_bytes(
        pd.DataFrame(
            [
                {
                    "生产": "1 KG",
                    "发票数量(kg)": "1 CNY/KG",
                    "销售单价": "1 CNY",
                },
                {
                    "生产": "CA1131021          麦乐鸡块 500g*30",
                    "发票数量(kg)": 3750,
                    "销售单价": 17.83,
                },
            ]
        )
    )


def sample_mapping() -> BytesIO:
    return to_excel_bytes(
        pd.DataFrame(
            [
                {"出口地": "新加坡", "售达方": "新加坡CMM", "产品名称": "新加坡CMM 脆皮鸡肉1000g/12袋/箱"},
                {"出口地": None, "售达方": None, "产品名称": "新加坡DC照烧鸡排100g/10枚/12袋/箱"},
            ]
        )
    )


def test_sample_transform_values():
    tables = process_files(sample_dalian(), sample_tieling())

    dalian_row = tables.dalian[tables.dalian["产品名称"] == "新加坡CMM 脆皮鸡肉1000g/12袋/箱"].iloc[0]
    assert dalian_row["月销量（吨）"] == 2.4
    assert dalian_row["单价(元/吨)"] == 3880
    assert dalian_row["币种"] == "USD"

    tieling_row = tables.tieling[tables.tieling["产品名称"] == "麦乐鸡块 500g*30"].iloc[0]
    assert tieling_row["月销量（吨）"] == 3.75
    assert round(tieling_row["单价(元/吨)"], 6) == 17830
    assert tieling_row["币种"] == "RMB"

    dc_extra = tables.dalian[tables.dalian["产品名称"] == "新加坡DC照烧鸡排100g/10枚/12袋/箱"].iloc[0]
    assert dc_extra["月销量（吨）"] == 1.56


def test_export_workbook_has_headers_formulas_and_merges(tmp_path):
    tables = process_files(sample_dalian(), sample_tieling())
    output = tmp_path / "summary.xlsx"
    output.write_bytes(export_summary_workbook(tables.final, month_label="5月"))

    wb = load_workbook(output, data_only=False)
    ws = wb.active

    assert ws["A1"].value == "外销市场"
    assert ws["D1"].value == "5月销量\n（吨）"
    assert ws["D4"].value == "=SUM(D2:D3)"
    assert ws["D6"].value == "=SUM(D5:D5)"
    assert "A2:A4" in {str(rng) for rng in ws.merged_cells.ranges}


def test_customer_mapping_can_fill_missing_dalian_customer():
    raw = pd.read_excel(sample_dalian(), dtype=object)
    raw["售达方.1"] = None

    mapping = load_customer_mapping(sample_mapping())
    dalian, unmapped = transform_dalian(to_excel_bytes(raw), mapping)

    row = dalian[dalian["产品名称"] == "新加坡CMM 脆皮鸡肉1000g/12袋/箱"].iloc[0]
    assert row["外销客户"] == "新加坡CMM"
    assert row["外销市场"] == "新加坡"
    assert unmapped.empty
