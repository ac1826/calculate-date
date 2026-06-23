from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import BinaryIO, Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_COLUMNS = ["外销市场", "外销客户", "产品名称", "月销量（吨）", "单价(元/吨)", "币种", "出口方"]

MARKET_ORDER = {"新加坡": 0, "香港": 1}
EXPORTER_ORDER = {"大连": 0, "铁岭": 1}

CUSTOMER_ORDER = {
    "新加坡CMM": 0,
    "COMMONWEALTH FOOD SERVICES PTE LTD": 1,
    "DOUBLE CHIN FOOD SERVICES PTE LTD": 2,
    "EASTERN HARVEST FOODS(SINGAPORE) PT": 3,
    "大成万达（香港）有限公司": 4,
}

PRODUCT_ORDER = {
    "新加坡CMM 脆皮鸡肉1000g/12袋/箱": 0,
    "新加坡CMM 紫菜脆皮鸡肉450g/20袋/箱": 1,
    "新加坡CMM 紫菜脆皮鸡肉1000g/12袋/箱": 2,
    "新加坡-CMM 唐扬（原味）450g/20袋/箱": 3,
    "新加坡CMM香脆鸡米花450g/20袋/箱": 4,
    "新加坡CMM黑胡椒鸡腿块450g/20袋/箱": 5,
    "新加坡GM 紫菜脆皮鸡肉1000g/6袋/箱": 6,
    "新加坡DC照烧鸡排100g/10枚/12袋/箱": 7,
    "新加坡DC100g炸鸡排10枚/12袋/箱": 8,
    "新加坡DC裹面包屑炸鸡排100g/10枚/12袋/箱": 9,
    "新加坡DC冷冻熟鸡丝1kg/10袋/箱": 10,
    "新加坡DC唐扬1kg/12袋/箱": 11,
    "新加坡DC紫菜脆皮鸡肉1kg/12袋/箱": 12,
    "新加坡DC鸡柳1kg/12袋/箱": 13,
    "新加坡-EH 紫菜脆皮鸡肉1000g/12袋/箱": 14,
    "新加坡EH27g炭烤腿肉串业务包27g/50串/6盒": 15,
    "新加坡EH27g炭烤腿肉串业务包2/50串/6盒/箱": 16,
    "香港大昌行日式烧汁鸡粒1kg/10袋/箱": 17,
    "香港-大昌行27g炭烤腿肉串2 10串/30袋/箱": 18,
}


@dataclass(frozen=True)
class ProcessedTables:
    dalian: pd.DataFrame
    tieling: pd.DataFrame
    final: pd.DataFrame
    unmapped: pd.DataFrame


def _read_excel(source: str | Path | BinaryIO) -> pd.DataFrame:
    return pd.read_excel(source, dtype=object)


def _number(value: object) -> float:
    if pd.isna(value):
        return 0.0
    if isinstance(value, str):
        value = value.replace(",", "").strip()
        if not value:
            return 0.0
    return float(value)


def _column_matches(columns: Iterable[str], name: str) -> list[str]:
    return [col for col in columns if col == name or col.startswith(f"{name}.")]


def _has_column(columns: Iterable[str], name: str) -> bool:
    return bool(_column_matches(columns, name))


def _pick_column(columns: Iterable[str], name: str, occurrence: int = 0) -> str:
    matches = _column_matches(columns, name)
    if len(matches) <= occurrence:
        raise ValueError(f"缺少字段: {name}")
    return matches[occurrence]


def _pick_first_column(columns: Iterable[str], names: Iterable[str], occurrence: int = 0) -> str:
    missing = []
    for name in names:
        try:
            return _pick_column(columns, name, occurrence)
        except ValueError:
            missing.append(name)
    columns = list(columns)
    if columns:
        return columns[0]
    raise ValueError(f"缺少字段: {' / '.join(missing)}")


def _market_from_text(*parts: object) -> str:
    text = " ".join("" if pd.isna(part) else str(part) for part in parts)
    return "香港" if "香港" in text else "新加坡"


def _clean_text(value: object) -> str:
    return "" if pd.isna(value) else str(value).strip()


def load_customer_mapping(source: str | Path | BinaryIO | None) -> pd.DataFrame:
    if source is None:
        return pd.DataFrame(columns=["出口地", "售达方", "产品名称"])

    raw = pd.read_excel(source, sheet_name=0, dtype=object)
    required = {"出口地", "售达方", "产品名称"}
    missing = required.difference(raw.columns)
    if missing:
        raise ValueError(f"清单缺少字段: {', '.join(sorted(missing))}")

    mapping = raw[["出口地", "售达方", "产品名称"]].copy()
    mapping[["出口地", "售达方"]] = mapping[["出口地", "售达方"]].ffill()
    for column in mapping.columns:
        mapping[column] = mapping[column].map(_clean_text)
    mapping = mapping[(mapping["产品名称"] != "") & (mapping["售达方"] != "")]
    return mapping.drop_duplicates(subset=["产品名称"], keep="first").reset_index(drop=True)


def _sort_key(row: pd.Series) -> tuple:
    return (
        MARKET_ORDER.get(row["外销市场"], 99),
        EXPORTER_ORDER.get(row["出口方"], 99),
        CUSTOMER_ORDER.get(row["外销客户"], 99),
        PRODUCT_ORDER.get(row["产品名称"], 999),
        row.get("_source_order", 999999),
        row["外销客户"],
        row["产品名称"],
    )


def _mapping_by_product(customer_mapping: pd.DataFrame | str | Path | BinaryIO | None) -> dict:
    mapping = (
        load_customer_mapping(customer_mapping)
        if not isinstance(customer_mapping, pd.DataFrame)
        else customer_mapping.copy()
    )
    return mapping.set_index("产品名称").to_dict("index") if not mapping.empty else {}


def transform_order_export(
    source: str | Path | BinaryIO,
    exporter: str,
    customer_mapping: pd.DataFrame | str | Path | BinaryIO | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    raw = _read_excel(source)
    columns = list(raw.columns)

    customer_col = _pick_column(columns, "售达方", 1) if _has_column(columns, "售达方") else None
    product_col = _pick_first_column(columns, ["描述", "产品名称", "品名", "物料描述", "产品描述"])
    quantity_col = _pick_column(columns, "订购数量", 0)
    price_col = _pick_column(columns, "单价PN00", 0)
    currency_col = _pick_column(columns, "单价PN00", 1)
    mapping_by_product = _mapping_by_product(customer_mapping)

    rows = []
    unmapped_rows = []
    for source_order, record in raw.iterrows():
        quantity_kg = _number(record[quantity_col])
        customer = _clean_text(record[customer_col]) if customer_col else ""
        product = _clean_text(record[product_col])
        currency = _clean_text(record[currency_col])
        price_per_kg = _number(record[price_col])
        market = _market_from_text(customer, product)

        if product and (not customer) and product in mapping_by_product:
            mapped = mapping_by_product[product]
            customer = mapped["售达方"]
            market = mapped["出口地"] or _market_from_text(customer, product)

        if not product:
            continue
        if not customer:
            unmapped_rows.append({"产品名称": product, "_source_order": source_order})
            continue

        rows.append(
            {
                "外销市场": market,
                "外销客户": customer,
                "产品名称": product,
                "quantity_kg": quantity_kg,
                "price_per_kg": price_per_kg,
                "币种": currency or "USD",
                "出口方": exporter,
                "_source_order": source_order,
            }
        )

    if not rows:
        return pd.DataFrame(columns=OUTPUT_COLUMNS + ["_source_order"]), pd.DataFrame(unmapped_rows).drop_duplicates()

    prepared = pd.DataFrame(rows)
    grouped = (
        prepared.groupby(["外销市场", "外销客户", "产品名称", "price_per_kg", "币种", "出口方"], as_index=False)
        .agg({"quantity_kg": "sum", "_source_order": "min"})
    )
    grouped = grouped[grouped["quantity_kg"] > 0].copy()
    grouped["月销量（吨）"] = grouped["quantity_kg"] / 1000
    grouped["单价(元/吨)"] = grouped["price_per_kg"] * 1000
    result = grouped[OUTPUT_COLUMNS + ["_source_order"]].copy()
    unmapped = pd.DataFrame(unmapped_rows).drop_duplicates() if unmapped_rows else pd.DataFrame(columns=["产品名称", "_source_order"])
    return result.sort_values(by="_source_order", kind="stable").reset_index(drop=True), unmapped


def transform_dalian(
    source: str | Path | BinaryIO,
    customer_mapping: pd.DataFrame | str | Path | BinaryIO | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    return transform_order_export(source, "大连", customer_mapping)


def _strip_product_code(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    return re.sub(r"^[A-Z0-9]+\s+", "", text).strip()


def transform_invoice_detail(
    source: str | Path | BinaryIO,
    exporter: str,
    customer_mapping: pd.DataFrame | str | Path | BinaryIO | None = None,
    default_market: str = "香港",
    default_customer: str = "大成万达（香港）有限公司",
) -> tuple[pd.DataFrame, pd.DataFrame]:
    raw = _read_excel(source)
    columns = list(raw.columns)
    product_col = _pick_first_column(columns, ["生产", "产品名称", "品名", "物料描述", "产品描述"], 0)
    quantity_col = _pick_first_column(columns, ["发票数量(kg)", "发票数量", "数量", "订购数量"], 0)
    price_col = _pick_first_column(columns, ["销售单价", "单价", "单价PN00"], 0)
    mapping_by_product = _mapping_by_product(customer_mapping)
    start_row = 0
    if not raw.empty:
        first_quantity = raw.iloc[0][quantity_col]
        first_price = raw.iloc[0][price_col]
        if any(ch.isalpha() for ch in str(first_quantity)) or any(ch.isalpha() for ch in str(first_price)):
            start_row = 1

    rows = []
    unmapped_rows = []
    for source_order, record in raw.iloc[start_row:].iterrows():
        product = _strip_product_code(record[product_col])
        quantity_kg = _number(record[quantity_col])
        price_per_kg = _number(record[price_col])
        if not product or quantity_kg == 0:
            continue

        market = default_market
        customer = default_customer
        if product in mapping_by_product:
            mapped = mapping_by_product[product]
            customer = mapped["售达方"] or customer
            market = mapped["出口地"] or market
        elif not default_customer:
            unmapped_rows.append({"产品名称": product, "_source_order": source_order})
            continue

        rows.append(
            {
                "外销市场": market,
                "外销客户": customer,
                "产品名称": product,
                "月销量（吨）": quantity_kg / 1000,
                "单价(元/吨)": price_per_kg * 1000,
                "币种": "RMB",
                "出口方": exporter,
                "_source_order": source_order,
            }
        )

    return pd.DataFrame(rows, columns=OUTPUT_COLUMNS + ["_source_order"]), pd.DataFrame(unmapped_rows).drop_duplicates()


def transform_tieling(source: str | Path | BinaryIO) -> pd.DataFrame:
    result, _ = transform_invoice_detail(source, "铁岭")
    return result


def transform_export_file(
    source: str | Path | BinaryIO,
    exporter: str,
    customer_mapping: pd.DataFrame | str | Path | BinaryIO | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    raw = _read_excel(source)
    columns = list(raw.columns)
    if _has_column(columns, "订购数量") and _has_column(columns, "单价PN00"):
        return transform_order_export(source, exporter, customer_mapping)
    if _has_column(columns, "发票数量(kg)") and _has_column(columns, "销售单价"):
        return transform_invoice_detail(source, exporter, customer_mapping)
    raise ValueError("无法识别文件格式：需要包含“订购数量/单价PN00”或“发票数量(kg)/销售单价”")


def build_final_table(dalian: pd.DataFrame, tieling: pd.DataFrame, include_subtotals: bool = True) -> pd.DataFrame:
    combined = pd.concat([dalian, tieling], ignore_index=True)
    if combined.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    combined = combined.copy()
    combined["_sort"] = combined.apply(_sort_key, axis=1)
    combined = combined.sort_values("_sort", kind="stable").drop(columns=["_sort"]).reset_index(drop=True)

    if not include_subtotals:
        return combined[OUTPUT_COLUMNS].copy()

    blocks = []
    for market, group in combined.groupby("外销市场", sort=False):
        blocks.append(group[OUTPUT_COLUMNS])
        subtotal = {col: "" for col in OUTPUT_COLUMNS}
        subtotal["产品名称"] = "小计"
        subtotal["月销量（吨）"] = group["月销量（吨）"].sum()
        blocks.append(pd.DataFrame([subtotal]))
    return pd.concat(blocks, ignore_index=True)


def process_files(
    dalian_source: str | Path | BinaryIO,
    tieling_source: str | Path | BinaryIO,
    customer_mapping_source: str | Path | BinaryIO | None = None,
) -> ProcessedTables:
    customer_mapping = load_customer_mapping(customer_mapping_source)
    dalian, dalian_unmapped = transform_export_file(dalian_source, "大连", customer_mapping)
    tieling, tieling_unmapped = transform_export_file(tieling_source, "铁岭", customer_mapping)
    unmapped = pd.concat([dalian_unmapped, tieling_unmapped], ignore_index=True).drop_duplicates()
    final = build_final_table(dalian, tieling, include_subtotals=True)
    return ProcessedTables(dalian=dalian[OUTPUT_COLUMNS], tieling=tieling[OUTPUT_COLUMNS], final=final, unmapped=unmapped)


def infer_month_label(*filenames: str | None) -> str:
    for filename in filenames:
        if not filename:
            continue
        match = re.search(r"(\d{1,2})月", filename)
        if match:
            return f"{int(match.group(1))}月"
        date_match = re.search(r"(?:^|[^\d])\d{4}[.\-_/](\d{1,2})[.\-_/]\d{1,2}(?:[^\d]|$)", filename)
        if date_match:
            return f"{int(date_match.group(1))}月"
    return "本月"


def export_summary_workbook(final: pd.DataFrame, month_label: str = "本月") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = OUTPUT_COLUMNS.copy()
    headers[3] = f"{month_label}销量\n（吨）"
    headers[4] = "单价\n(元/吨)"
    ws.append(headers)

    current_market = ""
    for _, record in final.iterrows():
        values = [record.get(col, "") for col in OUTPUT_COLUMNS]
        is_subtotal = str(record.get("产品名称", "")) == "小计"
        if values[0]:
            current_market = values[0]
        if is_subtotal:
            values[0] = current_market
            values[3] = ""
        ws.append(values)

    _replace_subtotal_formulas(ws)
    _style_worksheet(ws)
    _merge_repeated_cells(ws)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _replace_subtotal_formulas(ws) -> None:
    group_start = 2
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=3).value == "小计":
            ws.cell(row=row, column=4).value = f"=SUM(D{group_start}:D{row - 1})"
            group_start = row + 1


def _style_worksheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    subtotal_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 36
    widths = [14, 34, 46, 14, 14, 10, 12]

    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(OUTPUT_COLUMNS)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.column == 3:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = Font(bold=True)
            if ws.cell(row=cell.row, column=3).value == "小计":
                cell.fill = subtotal_fill
                cell.font = Font(bold=True)

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=4).number_format = "0.00"
        ws.cell(row=row, column=5).number_format = "#,##0"


def _merge_repeated_cells(ws) -> None:
    for column in (1, 2, 7):
        start = 2
        current = ws.cell(row=start, column=column).value
        for row in range(3, ws.max_row + 2):
            value = ws.cell(row=row, column=column).value if row <= ws.max_row else None
            boundary = value != current or ws.cell(row=row - 1, column=3).value == "小计"
            if boundary:
                end = row - 1
                if current not in (None, "") and end > start:
                    ws.merge_cells(start_row=start, start_column=column, end_row=end, end_column=column)
                start = row
                current = value
